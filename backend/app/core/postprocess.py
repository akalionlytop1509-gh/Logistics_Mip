"""
core/postprocess.py
===================
Post-processing: metrics, professional Excel export, and (legacy) matplotlib network.

Functions
---------
analyze_results(results_df, total_cost)     → dict of metrics
save_detailed_results(results_df, metrics)  → writes formatted Excel + JSON
plot_network(results_df, ...)               → matplotlib figure (legacy fallback)
"""

import os
import json
import math

import pandas as pd
import networkx as nx
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

try:
    from openpyxl import load_workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ── JSON serialiser ────────────────────────────────────────────────────────────
class _Encoder(json.JSONEncoder):
    def default(self, obj):
        if hasattr(obj, "tolist"):
            return obj.tolist()
        if hasattr(obj, "item"):
            return obj.item()
        return super().default(obj)


# ── Mode colours (for matplotlib legend) ──────────────────────────────────────
_MODE_COLOR = {
    "road":         "#868e96",
    "barge":        "#339af0",
    "rail":         "#ff922b",
    "sea":          "#845ef7",
}
_NODE_COLOR = {
    "supply": "#ff6b6b",
    "demand": "#4dabf7",
    "port":   "#51cf66",
    "hub":    "#ced4da",
}


# ═════════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ═════════════════════════════════════════════════════════════════════════════

def analyze_results(results_df: pd.DataFrame, total_cost: float) -> dict:
    """Return a dict of summary metrics from solver output."""
    if results_df is None or results_df.empty:
        return {"Status": "Infeasible"}

    mode_dist = results_df.groupby("Mode")["Flow"].sum().to_dict()
    mode_dist = {k: int(v) for k, v in mode_dist.items()}

    # Export vs domestic split
    export_flow   = 0
    domestic_flow = 0
    if "Is_Export" in results_df.columns:
        export_flow   = int(results_df.loc[results_df["Is_Export"] == 1, "Flow"].sum())
        domestic_flow = int(results_df.loc[results_df["Is_Export"] != 1, "Flow"].sum())

    return {
        "Total_Cost":      float(total_cost),
        "Total_Flow":      int(results_df["Flow"].sum()),
        "Route_Count":     int(len(results_df)),
        "Mode_Distribution": mode_dist,
        "Export_Flow":     export_flow,
        "Domestic_Flow":   domestic_flow,
    }


def save_detailed_results(
    results_df: pd.DataFrame,
    metrics: dict,
    folder: str = "results",
) -> None:
    """
    Save results to:
      results/detailed_flow.xlsx  — colourised Excel workbook
      results/metrics.json        — metrics dict
    """
    os.makedirs(folder, exist_ok=True)

    xlsx_path = os.path.join(folder, "detailed_flow.xlsx")
    json_path = os.path.join(folder, "metrics.json")

    # ── Write JSON ────────────────────────────────────────────────────────────
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(metrics, f, indent=4, cls=_Encoder, ensure_ascii=False)

    # ── Write Excel ───────────────────────────────────────────────────────────
    mode_dist_df = pd.DataFrame(
        list(metrics.get("Mode_Distribution", {}).items()),
        columns=["Mode", "Total_Flow_TEU"],
    )

    summary_data = {
        "Metric": ["Total Cost (VND)", "Total Flow (TEU)", "Active Routes",
                   "Export Flow (TEU)", "Domestic Flow (TEU)"],
        "Value":  [
            metrics.get("Total_Cost", 0),
            metrics.get("Total_Flow", 0),
            metrics.get("Route_Count", 0),
            metrics.get("Export_Flow", 0),
            metrics.get("Domestic_Flow", 0),
        ],
    }
    summary_df = pd.DataFrame(summary_data)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary",     index=False)
        results_df.to_excel(writer, sheet_name="Detailed Flow", index=False)
        mode_dist_df.to_excel(writer, sheet_name="Mode Share", index=False)

    # ── Apply styling if openpyxl available ──────────────────────────────────
    if _OPENPYXL:
        _style_workbook(xlsx_path, results_df)


def plot_network(
    results_df: pd.DataFrame,
    supply_map: dict = None,
    demand_map: dict = None,
    port_nodes: set = None,
):
    """
    Legacy matplotlib network plot (fallback when Plotly not used).
    Returns a matplotlib Figure or None.
    """
    if results_df is None or results_df.empty:
        return None

    supply_map = supply_map or {}
    demand_map = demand_map or {}
    port_nodes = port_nodes or set()

    G = nx.DiGraph()
    for _, row in results_df.iterrows():
        G.add_edge(row["From"], row["To"], weight=row["Flow"], mode=row["Mode"])

    fig, ax = plt.subplots(figsize=(12, 9))
    ax.set_facecolor("#f8f9fa")
    fig.patch.set_facecolor("#ffffff")

    pos = nx.spring_layout(G, seed=42)

    # Node colours
    node_colors = []
    for node in G.nodes():
        if node in port_nodes:
            node_colors.append(_NODE_COLOR["port"])
        elif node in supply_map and node not in demand_map:
            node_colors.append(_NODE_COLOR["supply"])
        elif node in demand_map and node not in supply_map:
            node_colors.append(_NODE_COLOR["demand"])
        else:
            node_colors.append(_NODE_COLOR["hub"])

    # Edge colours
    max_flow = max((d["weight"] for _, _, d in G.edges(data=True)), default=1)
    edge_colors = []
    edge_widths = []
    for _, _, d in G.edges(data=True):
        mode_key = d.get("mode", "").lower()
        edge_colors.append(_MODE_COLOR.get(mode_key, "#adb5bd"))
        edge_widths.append(max(0.8, min(6, 0.8 + 5.2 * d["weight"] / max_flow)))

    nx.draw_networkx_nodes(G, pos, node_color=node_colors, node_size=1800, ax=ax)
    nx.draw_networkx_labels(G, pos, font_size=9, font_weight="bold", ax=ax)
    nx.draw_networkx_edges(
        G, pos,
        edge_color=edge_colors,
        width=edge_widths,
        arrows=True,
        arrowsize=20,
        ax=ax,
    )

    # Legend
    patches = [
        mpatches.Patch(color=_NODE_COLOR["supply"], label="Supply"),
        mpatches.Patch(color=_NODE_COLOR["demand"], label="Demand"),
        mpatches.Patch(color=_NODE_COLOR["port"],   label="Port/Cảng"),
        mpatches.Patch(color=_NODE_COLOR["hub"],    label="Hub"),
    ]
    for mode_key, colour in _MODE_COLOR.items():
        patches.append(mpatches.Patch(color=colour, label=mode_key.title()))

    ax.legend(handles=patches, loc="upper left", fontsize=8, framealpha=0.9)
    ax.set_title("Optimized Logistics Network", fontsize=14, fontweight="bold")
    ax.axis("off")
    plt.tight_layout()
    return fig


# ═════════════════════════════════════════════════════════════════════════════
# INTERNAL: Excel styling
# ═════════════════════════════════════════════════════════════════════════════

def _style_workbook(path: str, results_df: pd.DataFrame) -> None:
    """Apply colour formatting to the generated Excel workbook."""
    wb = load_workbook(path)

    _HEADER_FILL  = PatternFill("solid", fgColor="1A1A2E")
    _HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
    _EXPORT_FILL  = PatternFill("solid", fgColor="D3F9D8")   # green
    _ALT_FILL     = PatternFill("solid", fgColor="F8F9FA")   # light gray
    _BORDER       = Border(
        left=Side(style="thin", color="DEE2E6"),
        right=Side(style="thin", color="DEE2E6"),
        top=Side(style="thin", color="DEE2E6"),
        bottom=Side(style="thin", color="DEE2E6"),
    )

    def _fmt_sheet(ws):
        # Header row
        for cell in ws[1]:
            cell.fill   = _HEADER_FILL
            cell.font   = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER
        ws.row_dimensions[1].height = 22

        # Data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            is_export_row = False
            for cell in row:
                cell.border    = _BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Detect export rows
                col_name = ws.cell(row=1, column=cell.column).value or ""
                if "export" in str(col_name).lower() and str(cell.value) == "1":
                    is_export_row = True

            if is_export_row:
                for cell in row:
                    cell.fill = _EXPORT_FILL
            elif row_idx % 2 == 0:
                for cell in row:
                    cell.fill = _ALT_FILL

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(
                (len(str(c.value or "")) for c in col), default=0
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    for sheet_name in wb.sheetnames:
        _fmt_sheet(wb[sheet_name])

    wb.save(path)
