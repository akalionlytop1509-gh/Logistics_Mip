import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def autofit_columns(ws, max_len_limit=50):
    """
    Tự động căn chỉnh độ rộng cột của worksheet dựa trên chiều dài nội dung.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if val_str.startswith('='):
                continue
            # Bỏ qua các hàng tiêu đề chính ở đầu trang
            if cell.row < 5:
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), max_len_limit)

def generate_scenario_excel(summary_df, scenario_flows_dict, currency="VND"):
    """
    Tạo tệp Excel đa trang (.xlsx) cao cấp gồm:
    1. Sheet 1: Tổng quan Kịch bản kèm Dashboard biểu đồ so sánh Chi phí & CO2.
    2. Sheet 2: So sánh Tuyến đường (Route Flows) giữa các kịch bản.
    3. Sheet 3: Phân bổ Phương thức (Modal Split) kèm biểu đồ Stacked Column.
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    # Định nghĩa phong cách chung (Design System)
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1B365D")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="595959")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11)
    bold_data_font = Font(name=font_family, size=11, bold=True)
    
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    stripe_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    optimal_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
    infeasible_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # soft red
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 1: TỔNG QUAN KỊCH BẢN
    # ─────────────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tổng quan Kịch bản"
    ws1.views.sheetView[0].showGridLines = True
    
    # Tiêu đề trang
    ws1["A2"] = "BÁO CÁO SO SÁNH KỊCH BẢN LOGISTICS"
    ws1["A2"].font = title_font
    ws1["A3"] = "Hệ thống Tối ưu hóa Mạng lưới Vận tải Multimodal — Enterprise Dashboard"
    ws1["A3"].font = subtitle_font
    
    # Tạo Header bảng tổng quan
    headers1 = [
        "Kịch bản", "Trạng thái", f"Chi phí Mục tiêu ({currency})", "Phát thải CO₂ (tấn)", 
        "Tỷ lệ Đường bộ", "Tổng Lưu lượng (TEU)", "Tuyến Hoạt động", 
        "Nội địa (TEU)", "Xuất khẩu (TEU)"
    ]
    
    start_row = 5
    for col_idx, header in enumerate(headers1, start=1):
        cell = ws1.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws1.row_dimensions[start_row].height = 28
    
    # Điền dữ liệu kịch bản
    for row_idx, (_, row) in enumerate(summary_df.iterrows(), start=start_row + 1):
        ws1.row_dimensions[row_idx].height = 22
        is_stripe = (row_idx % 2 == 0)
        row_fill = stripe_fill if is_stripe else PatternFill(fill_type=None)
        
        # Scenario Name
        c_sc = ws1.cell(row=row_idx, column=1, value=row.get("scenario"))
        c_sc.font = bold_data_font
        c_sc.alignment = Alignment(horizontal="left", vertical="center")
        c_sc.fill = row_fill
        c_sc.border = thin_border
        
        # Status
        status_val = str(row.get("status", "")).upper()
        c_st = ws1.cell(row=row_idx, column=2, value=status_val)
        c_st.font = bold_data_font
        c_st.alignment = Alignment(horizontal="center", vertical="center")
        c_st.border = thin_border
        if "OPTIMAL" in status_val or "SUCCESS" in status_val:
            c_st.fill = optimal_fill
        else:
            c_st.fill = infeasible_fill
            
        # Metrics
        co2_val = row.get("co2_total")
        if pd.isna(co2_val) or co2_val is None:
            co2_val = 0
            
        num_format_cost = '$#,##0.00' if currency == "USD" else '#,##0'
        metrics_mapping = [
            (3, row.get("objective_value", 0), num_format_cost),
            (4, co2_val, '#,##0'),
            (5, row.get("road_share_pct", 0) / 100.0, '0.0%'),
            (6, row.get("total_flow", 0), '#,##0'),
            (7, row.get("active_routes", 0), '#,##0'),
            (8, row.get("domestic_flow", 0), '#,##0'),
            (9, row.get("export_flow", 0), '#,##0'),
        ]
        
        for col_c, val, num_format in metrics_mapping:
            cell = ws1.cell(row=row_idx, column=col_c, value=float(val) if pd.notna(val) else 0)
            cell.font = data_font
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = thin_border
            cell.number_format = num_format
 
    # TÍCH HỢP CHARTS (DASHBOARD)
    # Chart 1: So sánh Chi phí Mục tiêu (Bar Chart)
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = f"So sánh Chi phí Mục tiêu ({currency})"
    chart1.y_axis.title = currency
    chart1.x_axis.title = "Kịch bản"
    chart1.width = 14
    chart1.height = 8.5
    
    data1 = Reference(ws1, min_col=3, min_row=start_row, max_row=start_row + len(summary_df))
    cats1 = Reference(ws1, min_col=1, min_row=start_row + 1, max_row=start_row + len(summary_df))
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.legend = None # Không cần legend do chỉ có 1 cột
    ws1.add_chart(chart1, "B14")
    
    # Chart 2: Phát thải CO2 (Bar Chart)
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 13
    chart2.title = "So sánh Phát thải CO₂ (Tấn)"
    chart2.y_axis.title = "Tấn CO₂"
    chart2.x_axis.title = "Kịch bản"
    chart2.width = 14
    chart2.height = 8.5
    
    data2 = Reference(ws1, min_col=4, min_row=start_row, max_row=start_row + len(summary_df))
    cats2 = Reference(ws1, min_col=1, min_row=start_row + 1, max_row=start_row + len(summary_df))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.legend = None
    ws1.add_chart(chart2, "F14")

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 2: SO SÁNH TUYẾN ĐƯỜNG (ROUTE FLOW COMPARISON)
    # ─────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="So sánh Tuyến đường")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A2"] = "BẢNG SO SÁNH CHI TIẾT DÒNG CHẢY LOGISTICS"
    ws2["A2"].font = title_font
    ws2["A3"] = "Phân tích lưu lượng vận chuyển (TEU) theo từng tuyến đường và phương thức giữa các kịch bản"
    ws2["A3"].font = subtitle_font
    
    # Thu thập và tổng hợp toàn bộ các tuyến đường phát sinh lưu lượng
    all_routes = []
    for sc_name, flows in scenario_flows_dict.items():
        for f in flows:
            flow_val = float(f.get("Flow") or 0)
            if flow_val > 0.001:
                all_routes.append({
                    "From": f.get("From"),
                    "To": f.get("To"),
                    "Mode": f.get("Mode"),
                    "Is_Export": int(f.get("Is_Export", 0))
                })
                
    if not all_routes:
        ws2["A5"] = "Không phát hiện dòng chảy nào đang hoạt động trong các kịch bản được chọn."
        ws2["A5"].font = bold_data_font
    else:
        # Tạo bảng phẳng duy nhất các tuyến đường
        df_routes = pd.DataFrame(all_routes).drop_duplicates()
        
        # Merge lưu lượng của từng kịch bản
        for sc_name, flows in scenario_flows_dict.items():
            df_temp = pd.DataFrame(flows)
            if not df_temp.empty:
                df_temp = df_temp[["From", "To", "Mode", "Is_Export", "Flow"]].copy()
                df_temp.rename(columns={"Flow": sc_name}, inplace=True)
                df_routes = pd.merge(df_routes, df_temp, on=["From", "To", "Mode", "Is_Export"], how="left")
            else:
                df_routes[sc_name] = 0.0
                
        # Điền 0 cho các kịch bản không đi qua tuyến đó
        sc_names = list(scenario_flows_dict.keys())
        df_routes[sc_names] = df_routes[sc_names].fillna(0.0)
        df_routes.sort_values(by=["From", "To", "Mode"], inplace=True)
        
        # Headers của sheet 2
        headers2 = ["Điểm đầu (From)", "Điểm cuối (To)", "Phương thức (Mode)", "Loại hình Flow"] + sc_names
        
        for col_idx, header in enumerate(headers2, start=1):
            cell = ws2.cell(row=5, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        ws2.row_dimensions[5].height = 26
        
        # Viết dữ liệu
        for r_idx, (_, row) in enumerate(df_routes.iterrows(), start=6):
            ws2.row_dimensions[r_idx].height = 20
            is_stripe = (r_idx % 2 == 0)
            row_fill = stripe_fill if is_stripe else PatternFill(fill_type=None)
            
            # Key columns
            c_from = ws2.cell(row=r_idx, column=1, value=row.get("From"))
            c_to = ws2.cell(row=r_idx, column=2, value=row.get("To"))
            c_mode = ws2.cell(row=r_idx, column=3, value=row.get("Mode"))
            
            # Is Export
            is_exp = "Xuất khẩu (Export)" if row.get("Is_Export") == 1 else "Nội địa (Domestic)"
            c_exp = ws2.cell(row=r_idx, column=4, value=is_exp)
            
            for col_c, cell_c in enumerate([c_from, c_to, c_mode, c_exp], start=1):
                cell_c.font = data_font
                cell_c.fill = row_fill
                cell_c.border = thin_border
                cell_c.alignment = Alignment(horizontal="left" if col_c != 3 else "center", vertical="center")
                
            # Scenario flows columns
            for sc_idx, sc_name in enumerate(sc_names, start=5):
                flow_val = float(row.get(sc_name, 0.0))
                cell = ws2.cell(row=r_idx, column=sc_idx, value=flow_val)
                cell.font = data_font
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.0'

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 3: PHÂN BỔ PHƯƠNG THỨC (MODAL SPLIT COMPARISON)
    # ─────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet(title="Phân bổ Phương thức")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3["A2"] = "CƠ CẤU VÀ PHÂN BỔ PHƯƠNG THỨC VẬN TẢI"
    ws3["A2"].font = title_font
    ws3["A3"] = "So sánh tổng lưu lượng phân bổ (TEU) theo từng hình thức qua các kịch bản"
    ws3["A3"].font = subtitle_font
    
    headers3 = ["Kịch bản", "Đường bộ (Road)", "Đường sắt (Rail)", "Đường thủy/Biển (Water)", "Tổng cộng"]
    for col_idx, header in enumerate(headers3, start=1):
        cell = ws3.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    ws3.row_dimensions[5].height = 26
    
    for row_idx, (_, row) in enumerate(summary_df.iterrows(), start=6):
        ws3.row_dimensions[row_idx].height = 22
        is_stripe = (row_idx % 2 == 0)
        row_fill = stripe_fill if is_stripe else PatternFill(fill_type=None)
        
        # Scenario Name
        c_sc = ws3.cell(row=row_idx, column=1, value=row.get("scenario"))
        c_sc.font = bold_data_font
        c_sc.fill = row_fill
        c_sc.border = thin_border
        c_sc.alignment = Alignment(horizontal="left", vertical="center")
        
        road = float(row.get("road_flow", 0))
        rail = float(row.get("rail_flow", 0))
        water = float(row.get("water_flow", 0))
        total = road + rail + water
        
        split_data = [
            (2, road),
            (3, rail),
            (4, water),
            (5, total),
        ]
        
        for col_c, val in split_data:
            cell = ws3.cell(row=row_idx, column=col_c, value=val)
            cell.font = bold_data_font if col_c == 5 else data_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'
            
    # Chart 3: Stacked Column Chart cho Modal Split
    chart3 = BarChart()
    chart3.type = "col"
    chart3.grouping = "stacked"
    chart3.overlap = 100
    chart3.title = "Cơ cấu Lưu lượng theo Phương thức (TEU)"
    chart3.y_axis.title = "TEU"
    chart3.x_axis.title = "Kịch bản"
    chart3.width = 15
    chart3.height = 9.5
    
    # Trỏ đến dữ liệu của các phương thức Road, Rail, Water (không chọn tổng cộng)
    data3 = Reference(ws3, min_col=2, max_col=4, min_row=5, max_row=5 + len(summary_df))
    cats3 = Reference(ws3, min_col=1, min_row=6, max_row=5 + len(summary_df))
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    ws3.add_chart(chart3, "B14")

    # Căn chỉnh kích thước cột tự động cho cả 3 trang
    autofit_columns(ws1)
    autofit_columns(ws2)
    autofit_columns(ws3)
    
    wb.save(output)
    output.seek(0)
    return output.getvalue()
